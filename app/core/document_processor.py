import sys
from pathlib import Path

if sys.platform == "win32":
    import win32com.client
    import pythoncom
from langchain_text_splitters.character import RecursiveCharacterTextSplitter
from langchain_community.document_loaders import PyPDFLoader, Docx2txtLoader, TextLoader
from langchain_core.documents import Document


class WordDocLoader:
    """使用 Word COM 解析旧版 .doc 文件（仅 Windows）"""

    def __init__(self, file_path: str):
        self.file_path = file_path

    def load(self) -> list[Document]:
        if sys.platform != "win32":
            raise RuntimeError("旧版 .doc 格式仅在 Windows 上支持，请转换为 .docx 后重试")
        pythoncom.CoInitialize()
        word = None
        doc = None
        try:
            word = win32com.client.Dispatch("Word.Application")
            word.Visible = False
            doc = word.Documents.Open(str(Path(self.file_path).resolve()))
            text = doc.Content.Text
            if not text.strip():
                return []
            return [Document(page_content=text)]
        finally:
            if doc:
                doc.Close(False)
            if word:
                word.Quit()
            pythoncom.CoUninitialize()


class ExcelLoader:
    """解析 xlsx 文件，每行转为一条 '列名: 值' 文本"""

    SUMMARY_KEYWORDS = {"总计", "合计", "小计", "汇总", "总计:", "合计:", "小计:"}

    def __init__(self, file_path: str):
        self.file_path = file_path

    def load(self) -> list[Document]:
        from openpyxl import load_workbook
        wb = load_workbook(self.file_path, read_only=True, data_only=True)
        docs = []
        for ws in wb.worksheets:
            rows_iter = ws.iter_rows(values_only=True)
            headers = None
            excel_row_num = 0  # Excel 物理行号
            for row in rows_iter:
                excel_row_num += 1  # 每行递增
                if headers is None:
                    headers = [str(h) if h is None or str(h).strip() == "" else str(h) for i, h in enumerate(row)]
                    # 确保所有列名唯一
                    seen = {}
                    unique_headers = []
                    for h in headers:
                        if h in seen:
                            seen[h] += 1
                            unique_headers.append(f"{h}_{seen[h]}")
                        else:
                            seen[h] = 0
                            unique_headers.append(h)
                    headers = unique_headers
                    continue
                # 跳过空行
                if all(c is None for c in row):
                    continue
                # 跳过汇总行（任意列包含"总计"/"合计"/"小计"等）
                row_text = " ".join(str(c).strip() for c in row if c is not None)
                if any(kw in row_text for kw in self.SUMMARY_KEYWORDS):
                    continue
                pairs = [f"行号: {excel_row_num}"]
                for header, value in zip(headers, row):
                    if value is not None:
                        pairs.append(f"{header}: {value}")
                    else:
                        pairs.append(f"{header}: (空)")
                if pairs:
                    doc = Document(
                        page_content="\n".join(pairs),
                        metadata={"sheet_name": ws.title, "row_number": excel_row_num},
                    )
                    docs.append(doc)
        wb.close()
        return docs


class DocumentProcessor:
    """文档处理器：解析不同格式文件并按策略切片"""

    LOADER_MAP = {
        ".pdf": PyPDFLoader,
        ".docx": Docx2txtLoader,
        ".doc": WordDocLoader,
        ".txt": TextLoader,
        ".md": TextLoader,
        ".xlsx": ExcelLoader,
    }

    def __init__(self, chunk_size: int = 1000, chunk_overlap: int = 200):
        self.splitter = RecursiveCharacterTextSplitter(
            chunk_size=chunk_size,
            chunk_overlap=chunk_overlap,
            separators=["\n\n", "\n", "。", "；", "，", " ", ""],
        )

    def process(self, file_path: str, document_id: int, file_name: str) -> list[dict]:
        """解析文档并按策略切片，返回 [{text, metadata}, ...]"""
        path = Path(file_path)
        ext = path.suffix.lower()

        import logging
        logging.getLogger(__name__).info(f"文件路径: {file_path}, 检测扩展名: {ext}")

        loader_cls = self.LOADER_MAP.get(ext)
        if loader_cls is None:
            raise ValueError(f"不支持的文件格式: {ext}")

        loader = loader_cls(str(path))
        docs = loader.load()

        # Excel 每行已是独立完整记录，跳过切片
        if ext == ".xlsx":
            chunks = docs
        else:
            chunks = self.splitter.split_documents(docs)

        results = []
        for i, chunk in enumerate(chunks):
            metadata = {
                "document_id": document_id,
                "file_name": file_name,
                "chunk_index": i,
                "source": file_name,
            }
            # Excel chunk 保留 sheet_name 和 row_number 元数据
            if hasattr(chunk, "metadata"):
                if "sheet_name" in chunk.metadata:
                    metadata["sheet_name"] = chunk.metadata["sheet_name"]
                if "row_number" in chunk.metadata:
                    metadata["row_number"] = chunk.metadata["row_number"]

            # 文件名前缀：使文件名中的学号、姓名、关键词可被检索
            text = f"[文件: {file_name}]\n{chunk.page_content}"
            results.append({
                "text": text,
                "metadata": metadata,
            })
        return results

